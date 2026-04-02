"""
exportar.py — Endpoints de exportación de lista
------------------------------------------------
Requiere sesión válida via AuthMiddleware.
Genera XLSX con openpyxl con estilos ControlAnime.
Genera JSON limpio para importar en otras apps.
"""

import io
import json
from datetime import datetime

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, StreamingResponse

from backend.database.exportar_queries import get_lista_exportable

router = APIRouter()


def _uid(request: Request) -> int:
    return request.state.usuario_id


# ══════════════════════════════════════════════════════════════════════════════
# PREVIEW (datos para mostrar en frontend)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/exportar/data")
def exportar_preview(request: Request, tipo: str = "todo"):
    datos = get_lista_exportable(_uid(request), tipo)
    return JSONResponse(status_code=200, content={
        "animes": datos,
        "total":  len(datos),
        "tipo":   tipo,
    })


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT JSON
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/exportar/json")
def exportar_json(request: Request, tipo: str = "todo"):
    datos = get_lista_exportable(_uid(request), tipo)

    # Limpiar campos internos del export
    export = []
    for a in datos:
        export.append({
            "titulo":             a["titulo"],
            "titulo_alternativo": a["titulo_alternativo"],
            "categoria":          a["tipo"],
            "rating":             a["rating"],
            "episodios":          a["episodios"],
            "episodios_vistos":   a["episodios_vistos"],
            "generos":            a["genres_es"] or a["genres"],
            "estado":             a["estado"],
            "anio":               a["anio"],
            "temporada":          a["temporada"],
            "estudio":            a["estudio"],
            "tipo":               a["anime_tipo"],
            "duracion_min":       a["duracion"],
            "fecha_inicio":       a["fecha_inicio"],
            "fecha_fin":          a["fecha_fin"],
            "agregado_en":        a["agregado_en"],
        })

    payload = json.dumps({
        "exportado_en": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        "total":        len(export),
        "filtro":       tipo,
        "animes":       export,
    }, ensure_ascii=False, indent=2)

    filename = f"controlanime_{tipo}_{datetime.utcnow().strftime('%Y%m%d')}.json"

    return StreamingResponse(
        io.BytesIO(payload.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT XLSX
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/exportar/xlsx")
def exportar_xlsx(request: Request, tipo: str = "todo", username: str = ""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse(status_code=500, content={"error": "openpyxl no instalado"})

    datos = get_lista_exportable(_uid(request), tipo)

    # Intentar obtener username si no viene en query
    if not username:
        try:
            from backend.database.connection import get_db
            with get_db() as conn:
                cur = conn.cursor()
                cur.execute("SELECT username FROM usuarios WHERE id = %s", (_uid(request),))
                row = cur.fetchone()
                username = row[0] if row and row[0] else ""
                cur.close()
        except Exception:
            username = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Lista"

    # ── Paleta (fiel al template) ──────────────────────
    C_ACCENT   = "F97316"   # naranja
    C_HEADER_L = "1A1A1A"   # fondo título izquierdo
    C_HEADER_R = "3A3A3A"   # fondo título derecho
    C_CAB_BG   = "D0D0D0"   # fondo cabecera
    C_PIE_BG   = "E8E8E8"   # fondo pie
    C_NUM      = "888888"   # color números
    C_DARK     = "000000"
    C_WHITE    = "FFFFFF"
    C_MUTED    = "555555"

    thin = Side(style="thin", color="000000")
    bot_border = Border(bottom=Border(bottom=Side(style="thin", color="000000")).bottom)

    fecha_hoy = datetime.utcnow().strftime("%d de %B de %Y")
    filtro_label = {"todo": "Todo", "visto": "Vistos", "pendiente": "Pendientes", "favorito": "Favoritos"}.get(tipo, tipo.title())

    # Contadores para el pie
    cnt_total      = len(datos)
    cnt_vistos     = sum(1 for a in datos if a.get("tipo") == "visto")
    cnt_pendientes = sum(1 for a in datos if a.get("tipo") == "pendiente")
    cnt_favoritos  = sum(1 for a in datos if a.get("tipo") == "favorito")

    # ── COLUMNAS: A No. | B Título | C Capítulos | D Ep.vistos | E Género | F Fecha ──
    # (añadimos D = Ep. vistos respecto al template original que era A-E)

    # Fila 1: encabezado
    # B1:C1 → CONTROL ANIME (naranja sobre oscuro)
    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value     = "CONTROL ANIME"
    c.font      = Font(name="Calibri", bold=True, size=16, color=C_ACCENT)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_L)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # D1:F1 → fecha + filtro (blanco sobre gris oscuro)
    ws.merge_cells("D1:F1")
    c = ws["D1"]
    c.value     = f"Lista exportada el {fecha_hoy}  ·  Filtro: {filtro_label}"
    c.font      = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_R)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2: cabecera de columnas
    cabeceras = [
        ("A2", "No.",           "center"),
        ("B2", "TITULO",        "center"),
        ("C2", "CAPÍTULOS",     "center"),
        ("D2", "EP. VISTOS",    "center"),
        ("E2", "GÉNERO",        "center"),
        ("F2", "FECHA AGREGADO","center"),
    ]
    for addr, val, h in cabeceras:
        c = ws[addr]
        c.value     = val
        c.font      = Font(name="Calibri", bold=True, size=11, color=C_DARK)
        c.fill      = PatternFill("solid", fgColor=C_CAB_BG)
        c.alignment = Alignment(horizontal=h, vertical="center")
        c.border    = Border(bottom=Side(style="thin", color=C_DARK))

    # Filas de datos (desde fila 3)
    for i, a in enumerate(datos, 1):
        r = 2 + i

        eps_total  = a.get("episodios")
        eps_vistos = a.get("episodios_vistos")
        genero_raw = a.get("genres_es") or a.get("genres") or ""
        primer_genero = genero_raw.split(",")[0].strip().lower() if genero_raw else "-"
        fecha_str  = a.get("agregado_en", "")
        # Formatear fecha DD/MM/YYYY
        if fecha_str and len(fecha_str) >= 10:
            try:
                from datetime import date
                d = date.fromisoformat(fecha_str[:10])
                fecha_fmt = d.strftime("%d/%m/%Y")
            except Exception:
                fecha_fmt = fecha_str[:10]
        else:
            fecha_fmt = "-"

        eps_str = f"{eps_total} episodios" if eps_total else "-"
        ep_vis_str = str(eps_vistos) if eps_vistos else "-"

        row_data = [
            (i,            "center", C_NUM,  False),
            (a.get("titulo") or "-", "left", None, False),
            (eps_str,      "center", None,   False),
            (ep_vis_str,   "center", None,   False),
            (primer_genero,"center", None,   False),
            (fecha_fmt,    "center", None,   False),
        ]
        cols = ["A", "B", "C", "D", "E", "F"]
        for col, (val, h, color, bold) in zip(cols, row_data):
            c = ws[f"{col}{r}"]
            c.value     = val
            c.font      = Font(name="Calibri", size=10,
                               color=color if color else "000000",
                               bold=bold)
            c.alignment = Alignment(horizontal=h, vertical="center")

    # Fila separadora vacía (fila N+3)
    sep_row = 2 + len(datos) + 1

    # ── PIE DE PÁGINA ─────────────────────────────────
    # Fiel al template: merges A:C para texto, D:E para stats
    pie_start = sep_row + 1

    pie_fill = PatternFill("solid", fgColor=C_PIE_BG)

    # A:C merge — "Lista generada con ControlAnime"
    ws.merge_cells(f"A{pie_start}:C{pie_start}")
    c = ws[f"A{pie_start}"]
    c.value     = "Lista generada con ControlAnime · Uso personal y gratuito"
    c.font      = Font(name="Calibri", size=10, color=C_MUTED)
    c.fill      = pie_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Stats (columnas D y F)
    stats = [
        ("TOTAL",      cnt_total),
        ("VISTOS",     cnt_vistos),
        ("PENDIENTES", cnt_pendientes),
        ("FAVORITOS",  cnt_favoritos),
    ]
    for j, (label, val) in enumerate(stats):
        r = pie_start + j
        # Rellenar A:C con fondo
        for col in ["A", "B", "C"]:
            ws[f"{col}{r}"].fill = pie_fill
        # Usuario en A(pie_start+1):C(pie_start+1)
        if j == 1:
            ws.merge_cells(f"A{r}:C{r}")
            c2 = ws[f"A{r}"]
            c2.value     = f"Usuario: {username}" if username else ""
            c2.font      = Font(name="Calibri", size=10, color=C_MUTED)
            c2.fill      = pie_fill
            c2.alignment = Alignment(horizontal="center", vertical="center")
        elif j > 1:
            ws.merge_cells(f"A{r}:C{r}")
            ws[f"A{r}"].fill = pie_fill

        # Label
        dl = ws[f"D{r}"]
        dl.value     = label
        dl.font      = Font(name="Calibri", bold=True, size=10, color=C_DARK)
        dl.fill      = pie_fill
        dl.alignment = Alignment(horizontal="left", vertical="center")

        # Valor (en columna F para mantener proporciones con 6 cols)
        fv = ws[f"F{r}"]
        fv.value     = val
        fv.font      = Font(name="Calibri", bold=True, size=10, color=C_ACCENT)
        fv.fill      = pie_fill
        fv.alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{r}"].fill = pie_fill  # rellenar E también

    # ── Anchos de columna ─────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    # ── Congelar cabecera ─────────────────────────────
    ws.freeze_panes = "A3"

    # ── Propiedades ───────────────────────────────────
    wb.properties.title   = "ControlAnime - Lista"
    wb.properties.creator = "ControlAnime"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"controlanime_{tipo}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
