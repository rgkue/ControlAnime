"""
importar.py — Endpoints de importación de listas
-------------------------------------------------
Soporta: JSON propio de CA, CSV/XLSX, TXT, XML de MAL.
Flujo en 2 pasos:
  1. POST /importar/parse  → detecta formato, parsea, hace matching → preview
  2. POST /importar/confirmar → inserta los animes confirmados

Requiere sesión válida via AuthMiddleware.
"""

import csv
import io
import json
import re
import xml.etree.ElementTree as ET
from typing import Any

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse

from backend.database.importar_queries import resolver_lista, importar_animes

router = APIRouter()

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB


def _uid(request: Request) -> int:
    return request.state.usuario_id


# ══════════════════════════════════════════════════════════════════════════════
# PARSERS POR FORMATO
# ══════════════════════════════════════════════════════════════════════════════

def parse_json(contenido: bytes) -> list[dict]:
    """
    Acepta:
    - Export propio de CA: { animes: [ { titulo, categoria, ... } ] }
    - Lista simple:         [ { titulo, ... } ]
    - AniList:              { lists: [ { entries: [ { media: { title: { romaji } } } ] } ] }
    """
    data = json.loads(contenido.decode("utf-8-sig"))
    items = []

    # Formato CA propio
    if isinstance(data, dict) and "animes" in data:
        for a in data["animes"]:
            titulo = a.get("titulo") or a.get("title") or ""
            tipo   = a.get("categoria") or a.get("tipo") or a.get("status") or ""
            eps    = a.get("episodios_vistos") or a.get("episodes_watched")
            if titulo:
                items.append({"titulo": titulo, "tipo": tipo, "episodios_vistos": eps})

    # AniList export
    elif isinstance(data, dict) and "lists" in data:
        for lst in data["lists"]:
            status = lst.get("name", "")
            for entry in lst.get("entries", []):
                media  = entry.get("media", {})
                titulo = (media.get("title") or {}).get("romaji") or \
                         (media.get("title") or {}).get("english") or \
                         (media.get("title") or {}).get("native") or ""
                eps    = entry.get("progress")
                if titulo:
                    items.append({"titulo": titulo, "tipo": status, "episodios_vistos": eps})

    # Lista simple de objetos
    elif isinstance(data, list):
        for a in data:
            if isinstance(a, dict):
                titulo = a.get("titulo") or a.get("title") or a.get("name") or ""
                tipo   = a.get("tipo") or a.get("status") or a.get("categoria") or ""
                eps    = a.get("episodios_vistos") or a.get("episodes_watched")
                if titulo:
                    items.append({"titulo": titulo, "tipo": tipo, "episodios_vistos": eps})
            elif isinstance(a, str) and a.strip():
                items.append({"titulo": a.strip(), "tipo": "", "episodios_vistos": None})

    return items


def parse_txt(contenido: bytes) -> list[dict]:
    """
    Texto plano: un título por línea.
    Ignora líneas vacías y comentarios con #.
    """
    texto = contenido.decode("utf-8-sig", errors="replace")
    items = []
    for linea in texto.splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#"):
            continue
        items.append({"titulo": linea, "tipo": "", "episodios_vistos": None})
    return items


def parse_csv(contenido: bytes) -> list[dict]:
    """
    CSV con cabeceras flexibles.
    Detecta columnas: titulo/title/name, tipo/status/categoria, episodios_vistos/episodes.
    """
    texto  = contenido.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(texto))
    items  = []

    # Mapeo flexible de nombres de columna
    COL_TITULO = ("titulo", "title", "name", "anime", "serie", "series_title")
    COL_TIPO   = ("tipo", "status", "categoria", "category", "estado")
    COL_EPS    = ("episodios_vistos", "episodes_watched", "ep_vistos", "eps", "progress")

    def find_col(row: dict, opciones: tuple) -> str | None:
        keys_lower = {k.lower().strip(): k for k in row}
        for opt in opciones:
            if opt in keys_lower:
                return keys_lower[opt]
        return None

    for row in reader:
        col_t = find_col(row, COL_TITULO)
        col_s = find_col(row, COL_TIPO)
        col_e = find_col(row, COL_EPS)

        titulo = row.get(col_t, "").strip() if col_t else ""
        tipo   = row.get(col_s, "").strip() if col_s else ""
        eps    = row.get(col_e, "").strip() if col_e else None

        if titulo:
            items.append({
                "titulo":            titulo,
                "tipo":              tipo,
                "episodios_vistos":  eps if eps else None,
            })
    return items


def parse_xlsx(contenido: bytes) -> list[dict]:
    """
    XLSX: detecta automáticamente columnas por cabecera.
    Compatible con el export propio de CA y con formatos genéricos.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl no instalado")

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    items = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Buscar fila de cabeceras (primera fila no vacía con texto)
    header_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).lower().strip() for v in row if v is not None]
        # Detectar si esta fila parece una cabecera
        if any(v in ("titulo", "title", "name", "serie") for v in vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        # Sin cabecera reconocible: asumir col B como título (formato CA)
        for row in rows[2:]:  # Saltar fila de título y cabecera
            if row and len(row) > 1 and row[1]:
                titulo = str(row[1]).strip()
                eps    = str(row[3]).strip() if len(row) > 3 and row[3] and str(row[3]) != "-" else None
                if titulo and titulo not in ("TITULO", "No.", "Lista generada"):
                    items.append({"titulo": titulo, "tipo": "visto", "episodios_vistos": eps})
        return items

    headers = [str(v).lower().strip() if v else "" for v in rows[header_row_idx]]

    COL_TITULO = ("titulo", "title", "name", "serie", "series_title")
    COL_TIPO   = ("tipo", "status", "categoria", "category", "estado")
    COL_EPS    = ("ep. vistos", "episodios_vistos", "episodes_watched", "ep_vistos", "progress")

    def find_idx(headers, opts):
        for opt in opts:
            for i, h in enumerate(headers):
                if opt in h:
                    return i
        return None

    idx_t = find_idx(headers, COL_TITULO)
    idx_s = find_idx(headers, COL_TIPO)
    idx_e = find_idx(headers, COL_EPS)

    if idx_t is None:
        return []

    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= idx_t:
            continue
        titulo = str(row[idx_t]).strip() if row[idx_t] else ""
        tipo   = str(row[idx_s]).strip() if idx_s is not None and len(row) > idx_s and row[idx_s] else ""
        eps    = str(row[idx_e]).strip() if idx_e is not None and len(row) > idx_e and row[idx_e] else None

        if titulo and titulo != "None":
            items.append({
                "titulo":           titulo,
                "tipo":             tipo,
                "episodios_vistos": eps if eps and eps != "-" else None,
            })

    return items


def parse_xml_mal(contenido: bytes) -> list[dict]:
    """
    Parsea el XML estándar de MyAnimeList:
      <myanimelist>
        <anime>
          <series_title>...</series_title>
          <my_status>Completed</my_status>
          <my_watched_episodes>12</my_watched_episodes>
        </anime>
      </myanimelist>

    También acepta el formato hianime/xml2mal:
      <list><folder><name>Completed</name><data><item><name>...</name></item></data></folder></list>
    """
    root = ET.fromstring(contenido.decode("utf-8-sig", errors="replace"))
    items = []

    # Formato MAL estándar
    for anime in root.findall("anime"):
        titulo = _xml_text(anime, "series_title") or _xml_text(anime, "series_title_transliterated")
        status = _xml_text(anime, "my_status") or ""
        eps    = _xml_text(anime, "my_watched_episodes")
        if titulo:
            items.append({
                "titulo":           titulo,
                "tipo":             status,
                "episodios_vistos": eps if eps and eps != "0" else None,
            })

    # Formato hianime/xml2mal (fallback)
    if not items:
        for folder in root.findall(".//folder"):
            status = _xml_text(folder, "name") or ""
            for item in folder.findall(".//item"):
                titulo = _xml_text(item, "name") or _xml_text(item, "title") or ""
                if titulo:
                    items.append({
                        "titulo":           titulo,
                        "tipo":             status,
                        "episodios_vistos": None,
                    })

    return items


def _xml_text(el, tag: str) -> str | None:
    child = el.find(tag)
    return child.text.strip() if child is not None and child.text else None


# ══════════════════════════════════════════════════════════════════════════════
# DETECTAR FORMATO
# ══════════════════════════════════════════════════════════════════════════════

def detectar_formato(filename: str, contenido: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return "xlsx"
    if ext == "csv":
        return "csv"
    if ext in ("json",):
        return "json"
    if ext in ("xml",):
        return "xml"
    if ext in ("txt", "text", ""):
        # Intentar detectar por contenido
        head = contenido[:200].decode("utf-8-sig", errors="ignore").strip()
        if head.startswith("{") or head.startswith("["):
            return "json"
        if head.startswith("<?xml") or head.startswith("<myanimelist") or head.startswith("<list"):
            return "xml"
        return "txt"
    return "txt"


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT 1: PARSE (subir archivo → preview)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/importar/parse")
async def importar_parse(request: Request, archivo: UploadFile = File(...)):
    contenido = await archivo.read()

    if len(contenido) > MAX_FILE_SIZE:
        return JSONResponse(status_code=400, content={"error": "Archivo demasiado grande (máx. 5MB)"})

    filename = archivo.filename or "archivo"

    try:
        formato = detectar_formato(filename, contenido)

        if formato == "json":
            items = parse_json(contenido)
        elif formato == "csv":
            items = parse_csv(contenido)
        elif formato == "xlsx":
            items = parse_xlsx(contenido)
        elif formato == "xml":
            items = parse_xml_mal(contenido)
        else:
            items = parse_txt(contenido)

    except Exception as e:
        return JSONResponse(status_code=400, content={
            "error": f"No se pudo leer el archivo: {str(e)}"
        })

    if not items:
        return JSONResponse(status_code=400, content={
            "error": "No se encontraron animes en el archivo. Verifica el formato."
        })

    # Limitar a 500 animes por importación
    items = items[:500]

    # Resolver matches contra la DB
    resultado = resolver_lista(items)

    return JSONResponse(status_code=200, content={
        "formato":         formato,
        "total_parseados": len(items),
        "encontrados":     resultado["encontrados"],
        "no_encontrados":  resultado["no_encontrados"],
    })


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT 2: CONFIRMAR (insertar en lista_animes)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/importar/confirmar")
async def importar_confirmar(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "JSON inválido"})

    animes = body.get("animes", [])
    if not animes:
        return JSONResponse(status_code=400, content={"error": "No hay animes para importar"})

    if len(animes) > 500:
        return JSONResponse(status_code=400, content={"error": "Máximo 500 animes por importación"})

    resultado = importar_animes(_uid(request), animes)

    return JSONResponse(status_code=200, content={
        "mensaje":    f"{resultado['importados']} animes importados correctamente",
        "importados": resultado["importados"],
        "errores":    resultado["errores"],
    })
