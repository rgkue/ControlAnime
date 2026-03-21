"""
lista.py — Endpoints de lista de animes, likes, reseñas y exportación
-----------------------------------------------------------------------
Fase 5: La autenticación ya viene validada por AuthMiddleware.
request.state.usuario_id contiene el ID del usuario autenticado.

Refactor: toda la lógica SQL migrada a backend/database/lista_queries.py.
El router solo coordina: valida input → llama query → devuelve JSONResponse.
"""

from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, StreamingResponse

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from backend.database.lista_queries import (
    agregar_anime,
    eliminar_anime,
    get_lista,
    borrar_lista,
    eliminar_cuenta,
    get_lista_export,
    get_estado_anime,
    dar_like,
    quitar_like,
    get_resenas,
    upsert_resena,
    eliminar_resena,
    get_detalle_lista,
    actualizar_detalle_lista,
)
from backend.database.connection import invalidar_sesion

router = APIRouter()


def _uid(request: Request) -> int:
    return request.state.usuario_id


# ══════════════════════════════════════════════════════════════════════════════
# LISTA DE ANIMES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/lista/agregar")
def agregar_a_lista(request: Request, datos: dict):
    anime_id         = datos.get("anime_id")
    tipo             = datos.get("tipo")
    episodios_vistos = datos.get("episodios_vistos")
    fecha_inicio     = datos.get("fecha_inicio")
    fecha_fin        = datos.get("fecha_fin")

    if not anime_id or tipo not in ("visto", "pendiente", "abandonado"):
        return JSONResponse(status_code=400, content={"error": "Datos inválidos"})

    agregado_en = agregar_anime(
        _uid(request), anime_id, tipo,
        episodios_vistos, fecha_inicio, fecha_fin,
    )
    if agregado_en is None:
        return JSONResponse(status_code=500, content={"error": "Error interno"})

    return JSONResponse(status_code=200, content={
        "mensaje":     "Agregado correctamente",
        "agregado_en": agregado_en,
    })


@router.post("/lista/eliminar")
def eliminar_de_lista(request: Request, datos: dict):
    anime_id = datos.get("anime_id")
    if not anime_id:
        return JSONResponse(status_code=400, content={"error": "Falta anime_id"})

    ok = eliminar_anime(_uid(request), anime_id)
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})

    return JSONResponse(status_code=200, content={"mensaje": "Eliminado correctamente"})


@router.get("/lista")
def obtener_lista(request: Request):
    data = get_lista(_uid(request))
    return JSONResponse(status_code=200, content=data)


@router.delete("/lista/borrar-todo")
def borrar_toda_lista(request: Request):
    ok = borrar_lista(_uid(request))
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})
    return JSONResponse(status_code=200, content={"mensaje": "Lista borrada correctamente"})


# ══════════════════════════════════════════════════════════════════════════════
# ESTADO DE ANIME (para la ficha /dashboard/anime/:id)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/lista/estado/{anime_id}")
def estado_anime(request: Request, anime_id: str):
    if not anime_id or len(anime_id) > 64:
        return JSONResponse(status_code=400, content={"error": "ID inválido"})
    data = get_estado_anime(_uid(request), anime_id)
    return JSONResponse(status_code=200, content=data)


# ══════════════════════════════════════════════════════════════════════════════
# DETALLE DE ENTRADA EN LISTA (/mi-lista/anime/:id)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/lista/detalle/{anime_id}")
def detalle_lista(request: Request, anime_id: str):
    data = get_detalle_lista(_uid(request), anime_id)
    if data is None:
        return JSONResponse(status_code=404, content={"error": "No está en tu lista"})
    return JSONResponse(status_code=200, content=data)


@router.patch("/lista/detalle/{anime_id}")
def actualizar_detalle(request: Request, anime_id: str, datos: dict):
    _PERMITIDOS = {"tipo", "episodios_vistos", "fecha_inicio", "fecha_fin", "agregado_en"}
    campos = {k: v for k, v in datos.items() if k in _PERMITIDOS}
    if not campos:
        return JSONResponse(status_code=400, content={"error": "Sin campos válidos"})

    ok = actualizar_detalle_lista(_uid(request), anime_id, campos)
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})
    return JSONResponse(status_code=200, content={"mensaje": "Actualizado"})


# ══════════════════════════════════════════════════════════════════════════════
# LIKES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/anime-likes")
def toggle_like_on(request: Request, datos: dict):
    anime_id = datos.get("anime_id")
    if not anime_id or len(str(anime_id)) > 64:
        return JSONResponse(status_code=400, content={"error": "ID inválido"})

    ok = dar_like(_uid(request), anime_id)
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})
    return JSONResponse(status_code=200, content={"mensaje": "Like guardado"})


@router.delete("/anime-likes")
def toggle_like_off(request: Request, datos: dict):
    anime_id = datos.get("anime_id")
    if not anime_id:
        return JSONResponse(status_code=400, content={"error": "Falta anime_id"})

    ok = quitar_like(_uid(request), anime_id)
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})
    return JSONResponse(status_code=200, content={"mensaje": "Like quitado"})


# ══════════════════════════════════════════════════════════════════════════════
# RESEÑAS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/resenas/{anime_id}")
def obtener_resenas(request: Request, anime_id: str):
    if not anime_id or len(anime_id) > 64:
        return JSONResponse(status_code=400, content={"error": "ID inválido"})
    resenas = get_resenas(anime_id, _uid(request))
    return JSONResponse(status_code=200, content={"resenas": resenas})


@router.post("/resenas")
@router.put("/resenas")
def crear_o_actualizar_resena(request: Request, datos: dict):
    anime_id   = datos.get("anime_id")
    rating     = datos.get("rating")
    comentario = (datos.get("comentario") or "").strip()[:600]

    if not anime_id or not isinstance(rating, int) or not (1 <= rating <= 10):
        return JSONResponse(status_code=400, content={"error": "Datos inválidos"})

    resultado = upsert_resena(_uid(request), anime_id, rating, comentario or None)
    if resultado is None:
        return JSONResponse(status_code=500, content={"error": "Error interno"})

    return JSONResponse(status_code=200, content={
        "mensaje":   "Reseña guardada",
        "id":        resultado["id"],
        "creado_en": resultado["creado_en"],
    })


@router.delete("/resenas/{resena_id}")
def borrar_resena(request: Request, resena_id: int):
    eliminada = eliminar_resena(resena_id, _uid(request))
    if not eliminada:
        return JSONResponse(status_code=404, content={"error": "Reseña no encontrada"})
    return JSONResponse(status_code=200, content={"mensaje": "Reseña eliminada"})


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/export/json")
def exportar_json(request: Request, filtro: str = "todo"):
    lista = get_lista_export(_uid(request), filtro)
    payload = {
        "exportado_en": datetime.now().isoformat(),
        "usuario_id":   _uid(request),
        "filtro":       filtro,
        "total":        len(lista),
        "animes":       lista,
    }
    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    nombre = f"controlanime_{datetime.now().strftime('%Y-%m-%d')}.json"
    return StreamingResponse(
        BytesIO(data),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/export/xlsx")
def exportar_xlsx(request: Request, filtro: str = "todo"):
    lista = get_lista_export(_uid(request), filtro)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Lista"

    orange      = PatternFill("solid", fgColor="F97316")
    dark        = PatternFill("solid", fgColor="1A1A1A")
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_font  = Font(name="Arial", size=15, bold=True, color="FFFFFF")

    fecha_txt = datetime.now().strftime("%d de %B de %Y")
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws["B1"] = "CONTROL ANIME"
    ws["D1"] = f"Lista exportada el {fecha_txt}"
    ws["B1"].fill = orange
    ws["D1"].fill = dark
    ws["B1"].font = title_font
    ws["D1"].font = Font(name="Arial", size=12, bold=False, color="FFFFFF")
    ws["B1"].alignment = ws["D1"].alignment = Alignment(horizontal="center")

    headers = ["No.", "TÍTULO", "CAPÍTULOS", "GÉNERO", "FECHA AGREGADO"]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=11, bold=True)

    for i, anime in enumerate(lista, start=3):
        ws.cell(i, 1, i - 2)
        ws.cell(i, 2, anime["titulo"])
        ws.cell(i, 3, f"{anime['episodios']} episodios" if anime["episodios"] else "-")
        genres = (anime.get("genres") or "").split(",")
        ws.cell(i, 4, ", ".join(g.strip() for g in genres[:2] if g.strip()) or "-")
        ws.cell(i, 5, (anime["agregado_en"] or "")[:10])

    footer_row  = len(lista) + 4
    ws.merge_cells(
        start_row=footer_row, start_column=1,
        end_row=footer_row,   end_column=3,
    )
    ws.cell(footer_row, 1, "Lista generada con ControlAnime · Uso personal y gratuito")

    vistos    = len({str(x["id"]) for x in lista if x["tipo"] in ("visto", "favorito")})
    pendientes = len([x for x in lista if x["tipo"] == "pendiente"])
    favoritos  = len([x for x in lista if x["tipo"] == "favorito"])
    stats = [
        ("TOTAL",      len(lista)),
        ("VISTOS",     vistos),
        ("PENDIENTES", pendientes),
        ("FAVORITOS",  favoritos),
    ]
    for idx, (label, value) in enumerate(stats):
        ws.cell(footer_row + idx, 4, label)
        ws.cell(footer_row + idx, 5, value)

    ws.cell(footer_row + len(stats) + 1, 1, f"Usuario: {_uid(request)}")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    nombre = f"controlanime_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# CUENTA
# ══════════════════════════════════════════════════════════════════════════════

@router.delete("/cuenta/eliminar")
def eliminar_mi_cuenta(request: Request):
    usuario_id = _uid(request)
    ok = eliminar_cuenta(usuario_id)
    if not ok:
        return JSONResponse(status_code=500, content={"error": "Error interno"})

    response = JSONResponse(status_code=200, content={"mensaje": "Cuenta eliminada"})
    response.delete_cookie("session", path="/")
    return response
